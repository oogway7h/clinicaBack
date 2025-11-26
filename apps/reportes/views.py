import os
from django.shortcuts import render
from django.http import HttpResponse
from io import BytesIO
import io
import json
import zipfile
import traceback
from django.apps import apps
from django.db import connection

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER,TA_LEFT, TA_RIGHT
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from datetime import date, datetime,time,timedelta,timezone
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Count
from django.db.models.functions import TruncDate, TruncMonth
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side,PatternFill




try:
    from apps.historiasDiagnosticos.models import Paciente
except ImportError:
    print("Error: No se pudo importar el modelo Paciente. Usando Mock.")
    class Paciente: objects = type('obj', (object,), {'select_related': lambda *a, **k: Paciente.objects, 'all': lambda *a, **k: [], 'filter': lambda *a, **k: [], 'none': lambda *a, **k: []})()

try:
    from apps.cuentas.models import Usuario, Rol 
except ImportError:
    print("Error: No se pudo importar Usuario/Rol. Usando Mock.")
    class Usuario: objects = type('obj', (object,), {'select_related': lambda *a, **k: Usuario.objects, 'get': lambda *a, **k: None})()
    class Rol: pass

try:
    from apps.citas_pagos.models import Cita_Medica
except ImportError:
    print("Error: No se pudo importar el modelo Cita_Medica.")
    class Cita_Medica: objects = type('obj', (object,), {'select_related': lambda *a, **k: Cita_Medica.objects, 'all': lambda *a, **k: [], 'filter': lambda *a, **k: [], 'none': lambda *a, **k: []})()


try:
    from .nlp_service import procesar_comando_voz
except ImportError:
    print("ADVERTENCIA: No se pudo importar 'nlp_service'.")
    def procesar_comando_voz(texto):
        return {"error": "Servicio NLP no cargado."}

FONT_NAME = None 
def _get_optional_date_range(request):
    """ayuda a poner un rango de fechas en los reportes si no se dan pone todos
    """
    fecha_inicio = None
    fecha_fin = None
    try:
        fecha_inicio_str = request.query_params.get('fecha_inicio', None)
        fecha_fin_str = request.query_params.get('fecha_fin', None)
        

        if fecha_inicio_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        if fecha_fin_str:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            
    except (ValueError, TypeError):
        print(f"[WARN] Fechas inválidas recibidas para PDF: {fecha_inicio_str}, {fecha_fin_str}")
        pass 
    return fecha_inicio, fecha_fin



@api_view(['GET']) 
@permission_classes([IsAuthenticated])
def generar_reporte_pacientes_pdf(request):
    
    try:
        auth_user = request.user 
        usuario_perfil = Usuario.objects.select_related('rol', 'grupo').get(correo=auth_user.email)
        
        admin_grupo = usuario_perfil.grupo
        admin_rol = usuario_perfil.rol

    except Usuario.DoesNotExist:
        return HttpResponse(f"Error: Perfil de usuario no encontrado para {auth_user.email}.", status=403)
    except Exception as e:
        traceback.print_exc()
        return HttpResponse(f"Error obteniendo perfil de usuario: {e}", status=500)
    fecha_inicio, fecha_fin = _get_optional_date_range(request)
#filtrar pacientes por grupo
    try:
        pacientes_qs = Paciente.objects.select_related('usuario', 'usuario__grupo').order_by('usuario__nombre')
        
        if fecha_inicio and fecha_fin:
            
            pacientes_qs = pacientes_qs.filter(usuario__fecha_registro__date__range=[fecha_inicio, fecha_fin])
            print(f"[DEBUG] Filtrando PDF de Pacientes por fechas: {fecha_inicio} a {fecha_fin}")
        

        if admin_rol and admin_rol.nombre == 'superAdmin':
            pacientes_filtrados = pacientes_qs.all()
            titulo_reporte = "Listado General de Pacientes (Todos los Grupos)"
        elif admin_grupo:
            pacientes_filtrados = pacientes_qs.filter(usuario__grupo=admin_grupo)
            titulo_reporte = f"Listado de Pacientes - Clínica: {admin_grupo.nombre}"
        else:
            pacientes_filtrados = Paciente.objects.none()
            titulo_reporte = "Listado de Pacientes (Sin Grupo Asignado)"

        try:
            total = pacientes_filtrados.count()
        except Exception:
            total = len(list(pacientes_filtrados)) # Fallback para queries complejas

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error al consultar la base de datos (ver consola).", status=500)

    styles = getSampleStyleSheet()
    
    # Estilo para el Título
    title_style = ParagraphStyle('title', parent=styles['Heading1'], alignment=TA_CENTER, fontName=(FONT_NAME or 'Helvetica-Bold'))
    
    # Estilo para celdas normales (centradas)
    normal_center = ParagraphStyle('normal_center', parent=styles['Normal'], alignment=TA_CENTER, fontName=(FONT_NAME or 'Helvetica'))
    
    # Estilo para encabezados de tabla
    header_style = ParagraphStyle('header', parent=normal_center, fontName=(FONT_NAME or 'Helvetica-Bold'))

    normal_left = ParagraphStyle('normal_left', parent=styles['Normal'], 
                                 alignment=TA_LEFT, fontName=(FONT_NAME or 'Helvetica'))
    normal_right = ParagraphStyle('normal_right', parent=styles['Normal'], 
                                  alignment=TA_RIGHT, fontName=(FONT_NAME or 'Helvetica'), fontSize=9)
   

    data = [
        [
            Paragraph("<b>N° Historia Clínica</b>", header_style),
            Paragraph("<b>Nombre Completo</b>", header_style),
            Paragraph("<b>Correo Electrónico</b>", header_style),
            Paragraph("<b>ID Paciente</b>", header_style),
        ]
    ]

    try:
        for paciente in pacientes_filtrados: 
            numero = getattr(paciente, 'numero_historia_clinica', '')
            nombre = getattr(getattr(paciente, 'usuario', None), 'nombre', 'N/A')
            correo = getattr(getattr(paciente, 'usuario', None), 'correo', 'N/A')
            pid = getattr(paciente, 'id', 'N/A')
            
            data.append([
                Paragraph(str(numero), normal_center),
                Paragraph(str(nombre), normal_center),
                Paragraph(str(correo), normal_center),
                Paragraph(str(pid), normal_center),
            ])
    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error al preparar datos del PDF (ver consola).", status=500)

    try:
        if len(data) == 1: 
            data.append([Paragraph("No se encontraron pacientes registrados para este grupo.", normal_center), "", "", ""])
    except Exception:
        pass 

    #construye el pdf
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=72, bottomMargin=72,title=titulo_reporte)
        elements = []
        
        fecha_gen=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        elements.append(Paragraph(f"Generado el: {fecha_gen} por {usuario_perfil.nombre}", normal_right))
        elements.append(Spacer(1, 12))

        elements.append(Paragraph(titulo_reporte, title_style))
        elements.append(Spacer(1, 24))

        total_registros = len(data) - 1 # Restamos la fila de encabezado
        intro_texto = f"Este reporte detalla <b>{total_registros} paciente(s)</b>"
        if fecha_inicio and fecha_fin:
            intro_texto += f" registrado(s) entre las fechas <b>{fecha_inicio}</b> y <b>{fecha_fin}</b>."
        else:
            intro_texto += " (histórico completo)."
        elements.append(Paragraph(intro_texto, normal_left))
        elements.append(Spacer(1, 24))

        table = Table(data, colWidths=[120, 150, 180, 80], repeatRows=1)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#839DB8")), 
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#E6F7FF")), 
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ])
        table.setStyle(table_style)
        elements.append(table)

        doc.build(elements)
        
        buffer.seek(0)
        pdf_bytes = buffer.read()
        buffer.close()

        if len(pdf_bytes) == 0:
            return HttpResponse("Error: PDF generado está vacío. Revisa la consola.", status=500)
        
        head = pdf_bytes[:10]
        if not head.startswith(b'%PDF-'):
            print("[PDF] Error: El archivo no parece ser un PDF. Primeros bytes:", head)
            return HttpResponse("Error: El contenido generado no es un PDF válido. Revisa la consola.", status=500)

        
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="listado_pacientes_{admin_grupo.id if admin_grupo else "super"}.pdf"'
        response['Content-Length'] = str(len(pdf_bytes))
        return response

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error interno generando el PDF. Revisa la consola.", status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generar_reporte_medicos_pdf(request):
    
    #obetenerl el grupo
    try:
        auth_user = request.user
        usuario_perfil = Usuario.objects.select_related('rol', 'grupo').get(correo=auth_user.email)
        admin_grupo = usuario_perfil.grupo
        admin_rol = usuario_perfil.rol
    except Usuario.DoesNotExist:
        return HttpResponse(f"Error: Perfil de usuario no encontrado para {auth_user.email}.", status=403)
    except Exception as e:
        traceback.print_exc()
        return HttpResponse(f"Error obteniendo perfil de usuario: {e}", status=500)
    fecha_inicio, fecha_fin = _get_optional_date_range(request)
    #filtrar por grupo
    try:
        rol_medico = Rol.objects.get(nombre='medico')
        medicos_qs = Usuario.objects.select_related('grupo', 'rol').filter(rol=rol_medico).order_by('nombre')

        if fecha_inicio and fecha_fin:
            # El campo de fecha está en el mismo modelo Usuario
            medicos_qs = medicos_qs.filter(fecha_registro__date__range=[fecha_inicio, fecha_fin])
            print(f"[DEBUG] Filtrando PDF de Médicos por fechas: {fecha_inicio} a {fecha_fin}")
        

        if admin_rol and admin_rol.nombre == 'superAdmin':
            medicos_filtrados = medicos_qs.all()
            titulo_reporte = "Listado General de Médicos (Todos los Grupos)"
        elif admin_grupo:
            medicos_filtrados = medicos_qs.filter(grupo=admin_grupo)
            titulo_reporte = f"Listado de Médicos - Clínica: {admin_grupo.nombre}"
        else:
            medicos_filtrados = Usuario.objects.none()
            titulo_reporte = "Listado de Médicos (Sin Grupo Asignado)"

        try:
            total = medicos_filtrados.count()
        except Exception:
            total = len(list(medicos_filtrados))

    except Rol.DoesNotExist:
        return HttpResponse("Error: El Rol 'medico' no existe en la base de datos.", status=500)
    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error al consultar la base de datos (ver consola).", status=500)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Heading1'], alignment=TA_CENTER, fontName=(FONT_NAME or 'Helvetica-Bold'))
    normal_center = ParagraphStyle('normal_center', parent=styles['Normal'], alignment=TA_CENTER, fontName=(FONT_NAME or 'Helvetica'))
    header_style = ParagraphStyle('header', parent=normal_center, fontName=(FONT_NAME or 'Helvetica-Bold'))
    normal_left = ParagraphStyle('normal_left', parent=styles['Normal'], alignment=TA_LEFT, fontName=(FONT_NAME or 'Helvetica'))
    normal_right = ParagraphStyle('normal_right', parent=styles['Normal'], alignment=TA_RIGHT, fontName=(FONT_NAME or 'Helvetica'), fontSize=9)
    

    data = [
        [
            Paragraph("<b>Nombre Completo</b>", header_style),
            Paragraph("<b>Correo Electrónico</b>", header_style),
            Paragraph("<b>Teléfono</b>", header_style),
            Paragraph("<b>ID Médico</b>", header_style),
        ]
    ]

    try:
        for medico in medicos_filtrados: 
            nombre = getattr(medico, 'nombre', 'N/A')
            correo = getattr(medico, 'correo', 'N/A')
            telefono = getattr(medico, 'telefono', '') 
            mid = getattr(medico, 'id', 'N/A')
            
            data.append([
                Paragraph(str(nombre), normal_center),
                Paragraph(str(correo), normal_center),
                Paragraph(str(telefono), normal_center),
                Paragraph(str(mid), normal_center),
            ])
    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error al preparar datos del PDF (ver consola).", status=500)

    # Mensaje si no hay datos
    try:
        if len(data) == 1:
            data.append([Paragraph("No se encontraron médicos registrados para este grupo.", normal_center), "", "", ""])
    except Exception:
        pass

    #construccion del documento solo pdf 
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=72, bottomMargin=72,title=titulo_reporte)
        elements = []
        
        fecha_gen = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        elements.append(Paragraph(f"Generado el: {fecha_gen} por {usuario_perfil.nombre}", normal_right))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(titulo_reporte, title_style))
        elements.append(Spacer(1, 12))

        total_registros = len(data) - 1 # Restamos la fila de encabezado
        intro_texto = f"Este reporte detalla <b>{total_registros} médico(s)</b>"
        if fecha_inicio and fecha_fin:
            intro_texto += f" registrado(s) entre las fechas <b>{fecha_inicio}</b> y <b>{fecha_fin}</b>."
        else:
            intro_texto += " (histórico completo)."
        elements.append(Paragraph(intro_texto, normal_left))
        elements.append(Spacer(1, 24))

        elements.append(Spacer(1, 24))

        table = Table(data, colWidths=[180, 180, 80, 80], repeatRows=1)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#318666")), 
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#E6F7FF")),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ])
        table.setStyle(table_style)
        elements.append(table)

        doc.build(elements)
        
        buffer.seek(0)
        pdf_bytes = buffer.read()
        buffer.close()

        if len(pdf_bytes) == 0:
            return HttpResponse("Error: PDF generado está vacío. Revisa la consola.", status=500)
        
        head = pdf_bytes[:10]
        if not head.startswith(b'%PDF-'):
            print("[PDF] Error: El archivo no parece ser un PDF. Primeros bytes:", head)
            return HttpResponse("Error: El contenido generado no es un PDF válido. Revisa la consola.", status=500)

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        filename = f'listado_medicos_{admin_grupo.id if admin_grupo else "super"}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = str(len(pdf_bytes))
        return response

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error interno generando el PDF. Revisa la consola.", status=500)
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generar_reporte_citas_pdf(request):
    
    #obtener el grupo
    try:
        auth_user = request.user
        usuario_perfil = Usuario.objects.select_related('rol', 'grupo').get(correo=auth_user.email)
        admin_grupo = usuario_perfil.grupo
        admin_rol = usuario_perfil.rol
    except Usuario.DoesNotExist:
        return HttpResponse(f"Error: Perfil de usuario no encontrado para {auth_user.email}.", status=403)
    except Exception as e:
        traceback.print_exc()
        return HttpResponse(f"Error obteniendo perfil de usuario: {e}", status=500)
    fecha_inicio, fecha_fin = _get_optional_date_range(request)
    try:
        citas_qs = Cita_Medica.objects.select_related(
            'paciente__usuario',
            'grupo'
        ).order_by('-fecha', '-hora_inicio')

        if fecha_inicio and fecha_fin:
            # El campo de fecha está en el mismo modelo Cita_Medica
            citas_qs = citas_qs.filter(fecha__range=[fecha_inicio, fecha_fin])
            print(f"[DEBUG] Filtrando PDF de Citas por fechas: {fecha_inicio} a {fecha_fin}")
        

        if admin_rol and admin_rol.nombre == 'superAdmin':
            citas_filtradas = citas_qs.all()
            titulo_reporte = "Reporte General de Citas (Todos los Grupos)"
        elif admin_grupo:
            citas_filtradas = citas_qs.filter(grupo=admin_grupo)
            titulo_reporte = f"Reporte de Citas - Clínica: {admin_grupo.nombre}"
        else:
            citas_filtradas = Cita_Medica.objects.none()
            titulo_reporte = "Reporte de Citas (Sin Grupo Asignado)"

        try:
            total = citas_filtradas.count()
        except Exception:
            total = len(list(citas_filtradas))

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error al consultar la base de datos (ver consola).", status=500)


    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Heading1'], alignment=TA_CENTER, fontName=(FONT_NAME or 'Helvetica-Bold'))
    normal_center = ParagraphStyle('normal_center', parent=styles['Normal'], alignment=TA_CENTER, fontName=(FONT_NAME or 'Helvetica'))
    header_style = ParagraphStyle('header', parent=normal_center, fontName=(FONT_NAME or 'Helvetica-Bold'))
    normal_left = ParagraphStyle('normal_left', parent=styles['Normal'], alignment=TA_LEFT, fontName=(FONT_NAME or 'Helvetica'))
    normal_right = ParagraphStyle('normal_right', parent=styles['Normal'], alignment=TA_RIGHT, fontName=(FONT_NAME or 'Helvetica'), fontSize=9)
    
    data = [
        [
            Paragraph("<b>Fecha</b>", header_style),
            Paragraph("<b>Hora</b>", header_style),
            Paragraph("<b>Paciente</b>", header_style),
            Paragraph("<b>Estado</b>", header_style),
            Paragraph("<b>ID Cita</b>", header_style),
        ]
    ]

    try:
        for cita in citas_filtradas:
            fecha = getattr(cita, 'fecha', 'N/A')
            hora_inicio = getattr(cita, 'hora_inicio', time(0,0))
        
            paciente_nombre = getattr(getattr(getattr(cita, 'paciente', None), 'usuario', None), 'nombre', 'N/A')
            
            estado = cita.get_estado_cita_display() if hasattr(cita, 'get_estado_cita_display') else getattr(cita, 'estado_cita', 'N/A')
            
            cid = getattr(cita, 'id', 'N/A')
            
            data.append([
                Paragraph(str(fecha), normal_center),
                Paragraph(str(hora_inicio)[:5], normal_center), 
                Paragraph(str(paciente_nombre), normal_center),
                Paragraph(str(estado), normal_center),
                Paragraph(str(cid), normal_center),
            ])
    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error al preparar datos del PDF (ver consola).", status=500)

    
    try:
        if len(data) == 1:
            data.append([Paragraph("No se encontraron citas registradas para este grupo.", normal_center), "", "", "", ""])
    except Exception:
        pass

    
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=72, bottomMargin=72,title=titulo_reporte)
        elements = []

        fecha_gen = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        elements.append(Paragraph(f"Generado el: {fecha_gen} por {usuario_perfil.nombre}", normal_right))
        elements.append(Spacer(1, 12))
        
        elements.append(Paragraph(titulo_reporte, title_style))
        elements.append(Spacer(1, 12))

        total_registros = len(data) - 1 # Restamos la fila de encabezado
        intro_texto = f"Este reporte detalla <b>{total_registros} cita(s)</b>"
        if fecha_inicio and fecha_fin:
            intro_texto += f" agendada(s) entre las fechas <b>{fecha_inicio}</b> y <b>{fecha_fin}</b>."
        else:
            intro_texto += " (histórico completo)."
        elements.append(Paragraph(intro_texto, normal_left))
        elements.append(Spacer(1, 24))

        elements.append(Spacer(1, 24))

    
        table = Table(data, colWidths=[80, 60, 200, 100, 50], repeatRows=1)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#8AD0E8")), 
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#E6F7FF")), 
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ])
        table.setStyle(table_style)
        elements.append(table)

        doc.build(elements)
        
        buffer.seek(0)
        pdf_bytes = buffer.read()
        buffer.close()

        if len(pdf_bytes) == 0:
            return HttpResponse("Error: PDF generado está vacío. Revisa la consola.", status=500)
        
        head = pdf_bytes[:10]
        if not head.startswith(b'%PDF-'):
            print("[PDF] Error: El archivo no parece ser un PDF. Primeros bytes:", head)
            return HttpResponse("Error: El contenido generado no es un PDF válido. Revisa la consola.", status=500)

        
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        filename = f'reporte_citas_{admin_grupo.id if admin_grupo else "super"}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = str(len(pdf_bytes))
        return response

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error interno generando el PDF. Revisa la consola.", status=500)





@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reporte_citas_por_dia(request):
    
    #usuario 
    try:
        auth_user = request.user
        usuario_perfil = Usuario.objects.select_related('rol', 'grupo').get(correo=auth_user.email)
        admin_grupo = usuario_perfil.grupo
        admin_rol = usuario_perfil.rol
    except Usuario.DoesNotExist:
        return Response(
            {"error": f"Perfil de usuario no encontrado para {auth_user.email}."}, 
            status=status.HTTP_403_FORBIDDEN
        )
    except Exception as e:
        traceback.print_exc()
        return Response(
            {"error": f"Error obteniendo perfil de usuario: {e}"}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

    #obtener filtr por fechas
    try:
        #ultimos 30 dias si no se pone nada
        fecha_fin_dt = datetime.now().date()
        fecha_inicio_dt = fecha_fin_dt - timedelta(days=29)

        fecha_inicio_str = request.query_params.get('fecha_inicio', None)
        fecha_fin_str = request.query_params.get('fecha_fin', None)

        if fecha_inicio_str:
            fecha_inicio_dt = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        if fecha_fin_str:
            fecha_fin_dt = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
    
    except ValueError:
        return Response(
            {"error": "Formato de fecha inválido. Use AAAA-MM-DD."},
            status=status.HTTP_400_BAD_REQUEST
        )
    except Exception as e:
        return Response({"error": f"Error procesando fechas: {e}"}, status=status.HTTP_400_BAD_REQUEST)

    #filtrar datos por grupos
    try:
        citas_qs_base = Cita_Medica.objects.filter(
            fecha__range=[fecha_inicio_dt, fecha_fin_dt]
        )

        
        if admin_rol and admin_rol.nombre == 'superAdmin':
            citas_qs_filtradas = citas_qs_base
        elif admin_grupo:
            citas_qs_filtradas = citas_qs_base.filter(grupo=admin_grupo)
        else:
            citas_qs_filtradas = Cita_Medica.objects.none()

        data_agrupada = citas_qs_filtradas.annotate(
            dia=TruncDate('fecha')
        ).values(
            'dia'
        ).annotate(
            total=Count('id')
        ).order_by(
            'dia'
        )
        
        datos_grafico = [
            {
                "fecha": item['dia'].isoformat(), 
                "total": item['total']
            } 
            for item in data_agrupada
        ]

        #generrar los datos
        lista_citas_qs = citas_qs_filtradas.select_related(
            'paciente__usuario'
        ).order_by('-fecha', '-hora_inicio')[:25] # <--- LÍMITE DE 25

        lista_citas = []
        for cita in lista_citas_qs:
            paciente_nombre = getattr(getattr(getattr(cita, 'paciente', None), 'usuario', None), 'nombre', 'N/A')
            estado = cita.get_estado_cita_display() if hasattr(cita, 'get_estado_cita_display') else getattr(cita, 'estado_cita', 'N/A')
            
            lista_citas.append({
                "id": cita.id,
                "fecha": cita.fecha.isoformat(),
                "hora_inicio": cita.hora_inicio.strftime('%H:%M'), 
                "paciente": paciente_nombre,
                "estado": estado
            })
        
        
        response_data = {
            "datos_grafico": datos_grafico,
            "lista_citas": lista_citas
        }
        
        return Response(response_data, status=status.HTTP_200_OK)

    except Exception as e:
        traceback.print_exc()
        return Response(
            {"error": f"Error al consultar la base de datos: {e}"}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generar_reporte_citas_excel(request):
    
    #obtener el grupo por el usuario
    try:
        auth_user = request.user
        usuario_perfil = Usuario.objects.select_related('rol', 'grupo').get(correo=auth_user.email)
        admin_grupo = usuario_perfil.grupo
        admin_rol = usuario_perfil.rol
    except Usuario.DoesNotExist:
        return HttpResponse(f"Error: Perfil de usuario no encontrado.", status=403)
    except Exception as e:
        return HttpResponse(f"Error obteniendo perfil de usuario: {e}", status=500)

    #obtener filtros de fechas
    try:
        fecha_fin_dt = datetime.now().date()
        fecha_inicio_dt = fecha_fin_dt - timedelta(days=29)
        
        fecha_inicio_str = request.query_params.get('fecha_inicio', None)
        fecha_fin_str = request.query_params.get('fecha_fin', None)

        if fecha_inicio_str:
            fecha_inicio_dt = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        if fecha_fin_str:
            fecha_fin_dt = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            
    except ValueError:
        return HttpResponse("Formato de fecha inválido. Use AAAA-MM-DD.", status=400)
    except Exception as e:
        return HttpResponse(f"Error procesando fechas: {e}", status=400)

    #filtrar datos
    try:
        citas_qs_base = Cita_Medica.objects.filter(
            fecha__range=[fecha_inicio_dt, fecha_fin_dt]
        )

        if admin_rol and admin_rol.nombre == 'superAdmin':
            citas_filtrados = citas_qs_base.select_related(
                'paciente__usuario', 'grupo'
            ).order_by('-fecha', '-hora_inicio')
            titulo_reporte = f"Reporte Citas {fecha_inicio_dt} a {fecha_fin_dt} (Todos)"
        elif admin_grupo:
            citas_filtrados = citas_qs_base.filter(grupo=admin_grupo).select_related(
                'paciente__usuario'
            ).order_by('-fecha', '-hora_inicio')
            titulo_reporte = f"Reporte Citas {fecha_inicio_dt} a {fecha_fin_dt} ({admin_grupo.nombre})"
        else:
            citas_filtrados = Cita_Medica.objects.none()
            titulo_reporte = "Reporte Citas (Sin Grupo)"

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error al consultar la base de datos (ver consola).", status=500)

    #generar el archivo excel
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Citas"

        #estilos del archivo(es solo una tabla gaaaaaaaa)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="004A99", end_color="004A99", fill_type="solid")  
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
        #encabezados
        headers = ["ID Cita", "Fecha", "Hora Inicio", "Hora Fin", "Paciente", "Estado", "Notas"]
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # datos
        row_num = 2
        for cita in citas_filtrados:
            paciente_nombre = getattr(getattr(getattr(cita, 'paciente', None), 'usuario', None), 'nombre', 'N/A')
            estado = cita.get_estado_cita_display() if hasattr(cita, 'get_estado_cita_display') else getattr(cita, 'estado_cita', 'N/A')
            
            ws.append([
                cita.id,
                cita.fecha,
                cita.hora_inicio,
                cita.hora_fin,
                paciente_nombre,
                estado,
                getattr(cita, 'notas', '')
            ])
            
            #formatos
            ws[f'B{row_num}'].number_format = 'YYYY-MM-DD'
            ws[f'C{row_num}'].number_format = 'hh:mm'
            ws[f'D{row_num}'].number_format = 'hh:mm'
            row_num += 1
            
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter 
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f'reporte_citas_{admin_grupo.id if admin_grupo else "super"}_{fecha_inicio_dt}_a_{fecha_fin_dt}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error interno generando el archivo Excel. Revisa la consola.", status=500)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reporte_pacientes_por_mes_json(request):
    
    #obtener grupo
    try:
        auth_user = request.user 
        usuario_perfil = Usuario.objects.select_related('rol', 'grupo').get(correo=auth_user.email)
        admin_grupo = usuario_perfil.grupo
        admin_rol = usuario_perfil.rol
    except Usuario.DoesNotExist:
        return Response({"error": "Perfil de usuario no encontrado."}, status=status.HTTP_403_FORBIDDEN)
    except Exception as e:
        return Response({"error": f"Error obteniendo perfil: {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    #obtener fechas
    try:
        fecha_fin = datetime.now().date()
        fecha_inicio = fecha_fin - timedelta(days=29)

        fecha_inicio_str = request.query_params.get('fecha_inicio', None)
        fecha_fin_str = request.query_params.get('fecha_fin', None)

        if fecha_inicio_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        if fecha_fin_str:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
    except Exception as e:
        return Response({"error": f"Formato de fecha inválido: {e}"}, status=status.HTTP_400_BAD_REQUEST)

    #Filtrar Pacientes
    try:
        pacientes_qs = Paciente.objects.filter(
            usuario__fecha_registro__date__range=[fecha_inicio, fecha_fin]
        )

        if admin_rol and admin_rol.nombre == 'superAdmin':
            pacientes_filtrados = pacientes_qs
        elif admin_grupo:
            pacientes_filtrados = pacientes_qs.filter(usuario__grupo=admin_grupo)
        else:
            pacientes_filtrados = Paciente.objects.none() 

        datos_grafico = (
            pacientes_filtrados
            .annotate(mes_registro=TruncMonth('usuario__fecha_registro'))
            .values('mes_registro')
            .annotate(total=Count('id'))
            .order_by('mes_registro')
        )
        
        datos_grafico_formato = [
            {"mes": item['mes_registro'].strftime('%Y-%m'), "total": item['total']}
            for item in datos_grafico
        ]

        lista_pacientes_detalle = (
            pacientes_filtrados
            .select_related('usuario')
            .order_by('-usuario__fecha_registro')[:25] 
        )
        
        lista_pacientes_formato = [
            {
                "id": p.id,
                "fecha_registro": p.usuario.fecha_registro.strftime('%Y-%m-%d'),
                "nombre": p.usuario.nombre,
                "correo": p.usuario.correo,
                "historia_clinica": p.numero_historia_clinica,
            }
            for p in lista_pacientes_detalle
        ]
        
        return Response({
            "datos_grafico": datos_grafico_formato,
            "lista_pacientes": lista_pacientes_formato
        }, status=status.HTTP_200_OK)

    except Exception as e:
        traceback.print_exc()
        return Response({"error": "Error al consultar la BD (ver consola)."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#view para la construccion del excel
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generar_reporte_pacientes_excel(request):
    
    try:
        auth_user = request.user 
        usuario_perfil = Usuario.objects.select_related('rol', 'grupo').get(correo=auth_user.email)
        admin_grupo = usuario_perfil.grupo
        admin_rol = usuario_perfil.rol
    except Usuario.DoesNotExist:
        return HttpResponse("Error: Perfil de usuario no encontrado.", status=403)
    except Exception as e:
        return HttpResponse(f"Error obteniendo perfil: {e}", status=500)

    try:
        today = datetime.now().date()
        default_start = today.replace(month=1, day=1).isoformat()
        
        fecha_fin_str = request.query_params.get('fecha_fin', today.isoformat())
        fecha_inicio_str = request.query_params.get('fecha_inicio', default_start)
        
        fecha_inicio = datetime.fromisoformat(fecha_inicio_str).date()
        fecha_fin = datetime.fromisoformat(fecha_fin_str).date()
    except Exception as e:
        return HttpResponse(f"Formato de fecha inválido: {e}", status=400)

    
    try:
        pacientes_qs = Paciente.objects.filter(
            usuario__fecha_registro__date__range=[fecha_inicio, fecha_fin]
        )

        if admin_rol and admin_rol.nombre == 'superAdmin':
            pacientes_filtrados = pacientes_qs.select_related('usuario').order_by('-usuario__fecha_registro')
            titulo_reporte = "Reporte de Pacientes Nuevos (Todos los Grupos)"
        elif admin_grupo:
            pacientes_filtrados = pacientes_qs.filter(usuario__grupo=admin_grupo).select_related('usuario').order_by('-usuario__fecha_registro')
            titulo_reporte = f"Reporte de Pacientes Nuevos - Clínica: {admin_grupo.nombre}"
        else:
            pacientes_filtrados = Paciente.objects.none() 
            titulo_reporte = "Reporte de Pacientes Nuevos (Sin Grupo Asignado)"

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error al consultar la BD (ver consola).", status=500)

    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Pacientes Nuevos"
        #encabezado
        sheet.append(["ID Paciente", "N° Historia Clínica", "Nombre Completo", "Correo", "Teléfono", "Fecha Registro"])

        #datos
        for paciente in pacientes_filtrados:
            sheet.append([
                paciente.id,
                paciente.numero_historia_clinica,
                paciente.usuario.nombre,
                paciente.usuario.correo,
                paciente.usuario.telefono or '',
                paciente.usuario.fecha_registro.strftime('%Y-%m-%d %H:%M')
            ])
        
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        excel_bytes = buffer.read()
        buffer.close()

        response = HttpResponse(
            excel_bytes, 
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_pacientes_nuevos_{fecha_inicio_str}_a_{fecha_fin_str}.xlsx"'
        return response

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error interno generando el Excel. Revisa la consola.", status=500)
    

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def procesar_comando_voz_json(request):
    """
    Recibe un comando de texto lo procesa con NLP
    y devuelve una acción que el fron va a ejecutar    """
    texto_comando = request.data.get('texto_comando', None)
    
    if not texto_comando:
        return Response(
            {"error": "No se proporcionó 'texto_comando'."}, 
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        resultado_nlp = procesar_comando_voz(texto_comando)
        
        if "error" in resultado_nlp:
            return Response(resultado_nlp, status=status.HTTP_400_BAD_REQUEST)
        
        return Response(resultado_nlp, status=status.HTTP_200_OK)
        
    except Exception as e:
        traceback.print_exc()
        return Response(
            {"error": f"Error interno en el servidor NLP: {e}"}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    

# def download_backup_json_zip(request):
#     ts = datetime.now().strftime("%Y%m%d_%H%M%S")
#     buf = io.BytesIO()

#     with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
#         # 1) Datos de cada modelo en JSON (1 archivo por modelo)
#         for model in apps.get_models():
#             label = model._meta.label  # p.ej. "app.Modelo"
#             # Evitar cargar todo en memoria
#             json_buf = io.StringIO()
#             json_buf.write("[")
#             first = True
#             for row in model.objects.all().values().iterator(chunk_size=1000):
#                 if not first:
#                     json_buf.write(",")
#                 json_buf.write(json.dumps(row, default=str))
#                 first = False
#             json_buf.write("]")
#             zf.writestr(f"{label}.json", json_buf.getvalue())

#         # 2) Esquema simple (columnas y tipos)
#         with connection.cursor() as cur:
#             cur.execute("""
#                 SELECT table_name, column_name, data_type
#                 FROM information_schema.columns
#                 WHERE table_schema='public'
#                 ORDER BY table_name, ordinal_position
#             """)
#             rows = cur.fetchall()
#         schema_txt = "\n".join(f"{t}.{c}  -  {d}" for t, c, d in rows)
#         zf.writestr("schema.txt", schema_txt)

#         # 3) Metadatos tomados del .env
#         meta = {
#             "generated_at": ts,
#             "db_name": os.getenv("DB_NAME"),
#             "db_user": os.getenv("DB_USER"),
#             "db_host": os.getenv("DB_HOST"),
#             "db_port": os.getenv("DB_PORT"),
#             "engine": "postgresql (via Django ORM)",
#         }
#         zf.writestr("metadata.json", json.dumps(meta, indent=2))

#     buf.seek(0)
#     resp = HttpResponse(buf.getvalue(), content_type="application/zip")
#     resp["Content-Disposition"] = f'attachment; filename="backup_json_{ts}.zip"'
#     return resp

def download_backup_json_zip(request):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    buf = io.BytesIO()

    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        # 1) Datos de cada modelo en SQL (1 archivo por modelo)
        for model in apps.get_models():
            label = model._meta.label  # p.ej. "app.Modelo"
            table_name = model._meta.db_table
            
            sql_buf = io.StringIO()
            sql_buf.write(f"-- Backup de {label}\n")
            sql_buf.write(f"-- Tabla: {table_name}\n\n")
            
            # Obtener nombres de columnas
            fields = [f.column for f in model._meta.fields]
            columns_str = ", ".join(fields)
            
            # Generar INSERTs
            for row in model.objects.all().values().iterator(chunk_size=1000):
                values = []
                for field_name in fields:
                    value = row.get(field_name)
                    
                    # Formatear el valor según su tipo
                    if value is None:
                        values.append("NULL")
                    elif isinstance(value, bool):
                        values.append("TRUE" if value else "FALSE")
                    elif isinstance(value, (int, float)):
                        values.append(str(value))
                    elif isinstance(value, datetime):
                        values.append(f"'{value.isoformat()}'")
                    elif isinstance(value, date):
                        values.append(f"'{value.isoformat()}'")
                    else:
                        # Escapar comillas simples
                        escaped = str(value).replace("'", "''")
                        values.append(f"'{escaped}'")
                
                values_str = ", ".join(values)
                sql_buf.write(f"INSERT INTO {table_name} ({columns_str}) VALUES ({values_str});\n")
            
            sql_buf.write("\n")
            zf.writestr(f"{label}.sql", sql_buf.getvalue())

        # 2) Esquema completo (CREATE TABLE statements)
        with connection.cursor() as cur:
            cur.execute("""
                SELECT table_name
                FROM information_schema.tables
                WHERE table_schema='public' AND table_type='BASE TABLE'
                ORDER BY table_name
            """)
            tables = [row[0] for row in cur.fetchall()]
            
            schema_sql = io.StringIO()
            schema_sql.write("-- Esquema de la base de datos\n\n")
            
            for table in tables:
                cur.execute(f"""
                    SELECT column_name, data_type, character_maximum_length, 
                    is_nullable, column_default
                    FROM information_schema.columns
                    WHERE table_name = %s AND table_schema='public'
                    ORDER BY ordinal_position
                """, [table])
                
                schema_sql.write(f"-- Tabla: {table}\n")
                for col_name, dtype, max_len, nullable, default in cur.fetchall():
                    len_str = f"({max_len})" if max_len else ""
                    null_str = "NULL" if nullable == "YES" else "NOT NULL"
                    def_str = f" DEFAULT {default}" if default else ""
                    schema_sql.write(f"--   {col_name}: {dtype}{len_str} {null_str}{def_str}\n")
                schema_sql.write("\n")
            
            zf.writestr("schema.sql", schema_sql.getvalue())

        # 3) Metadatos
        meta = {
            "generated_at": ts,
            "db_name": os.getenv("DB_NAME"),
            "db_user": os.getenv("DB_USER"),
            "db_host": os.getenv("DB_HOST"),
            "db_port": os.getenv("DB_PORT"),
            "engine": "postgresql (via Django ORM)",
            "format": "SQL INSERT statements"
        }
        zf.writestr("metadata.json", json.dumps(meta, indent=2))

    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type="application/zip")
    resp["Content-Disposition"] = f'attachment; filename="backup_sql_{ts}.zip"'
    return resp